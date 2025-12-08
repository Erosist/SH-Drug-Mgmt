from __future__ import annotations

from collections import defaultdict
from datetime import datetime, timedelta
from io import BytesIO
from typing import Any, Dict, List, Tuple

from flask import (
    Blueprint,
    Response,
    current_app,
    jsonify,
    request,
    send_file,
)
from flask_jwt_extended import jwt_required
from sqlalchemy import and_, func

from audit import record_admin_action
from extensions import db
from models import CirculationRecord, InventoryItem, Order, Tenant, User

try:  # optional heavy deps; imported lazily below too
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.styles import getSampleStyleSheet
    from reportlab.pdfgen import canvas
    from reportlab.pdfbase import pdfmetrics
    from reportlab.pdfbase.cidfonts import UnicodeCIDFont
except Exception:  # pragma: no cover - if not installed, export_pdf will fail explicitly
    A4 = None  # type: ignore
    getSampleStyleSheet = None  # type: ignore
    canvas = None  # type: ignore
    pdfmetrics = None  # type: ignore
    UnicodeCIDFont = None  # type: ignore

try:
    from openpyxl import Workbook
except Exception:  # pragma: no cover
    Workbook = None  # type: ignore


bp = Blueprint("compliance", __name__, url_prefix="/api/compliance")


def _require_regulator(user: User | None) -> bool:
    return bool(user and user.role == "regulator")


def _get_current_user() -> User | None:
    # 复用 auth 模块中的逻辑会引入循环依赖，这里简单重查一次
    from flask_jwt_extended import get_jwt_identity

    identity = get_jwt_identity()
    try:
        user_id = int(identity)
    except (TypeError, ValueError):
        return None
    return User.query.get(user_id)


def _parse_iso_datetime(value: str | None) -> datetime | None:
    if not value:
        return None
    try:
        # 兼容带 Z 的 ISO 字符串
        cleaned = value.replace("Z", "+00:00")
        dt = datetime.fromisoformat(cleaned)
        if dt.tzinfo:
            dt = dt.replace(tzinfo=None) - (dt.utcoffset() or timedelta(0))
        return dt
    except Exception:
        return None


def _parse_range() -> Tuple[datetime, datetime]:
    """解析时间范围，默认最近30天。"""
    end_raw = request.args.get("end")
    start_raw = request.args.get("start")

    end_dt = _parse_iso_datetime(end_raw) or datetime.utcnow()
    start_dt = _parse_iso_datetime(start_raw) or (end_dt - timedelta(days=30))

    # 保证 start <= end
    if start_dt > end_dt:
        start_dt, end_dt = end_dt - timedelta(days=30), end_dt
    return start_dt, end_dt


def _apply_region_filter(query, region: str | None):
    if region:
        like = f"%{region}%"
        # 调用方通常已经 join 了 Tenant，这里只追加地址过滤
        query = query.filter(Tenant.address.ilike(like))
    return query


def _inventory_compliance(start: datetime, end: datetime, region: str | None) -> Dict[str, Any]:
    """库存合规分析：统计低库存、缺货、近效期、过期比例。"""
    # 仅统计“有库存责任”的主体：药店 / 供应商；物流公司不参与库存合规
    base_query = InventoryItem.query.join(Tenant, InventoryItem.tenant_id == Tenant.id).filter(
        Tenant.type.in_(["PHARMACY", "SUPPLIER"])
    )
    if region:
        base_query = _apply_region_filter(base_query, region)

    items: List[InventoryItem] = (
        base_query.options(
            db.joinedload(InventoryItem.tenant),
            db.joinedload(InventoryItem.drug),
        ).all()
    )

    per_tenant: Dict[int, Dict[str, Any]] = defaultdict(
        lambda: {
            "tenant_id": None,
            "tenant_name": "",
            "total_batches": 0,
            "low_stock": 0,
            "out_of_stock": 0,
            "near_expiry": 0,
            "expired": 0,
        }
    )

    for item in items:
        tid = item.tenant_id
        t_entry = per_tenant[tid]
        t_entry["tenant_id"] = tid
        t_entry["tenant_name"] = item.tenant.name if item.tenant else ""
        t_entry["total_batches"] += 1

        if item.quantity == 0:
            t_entry["out_of_stock"] += 1
        elif item.quantity < 10:
            t_entry["low_stock"] += 1

        days_to_expiry = item.days_to_expiry
        if days_to_expiry <= 0:
            t_entry["expired"] += 1
        elif days_to_expiry <= 30:
            t_entry["near_expiry"] += 1

    tenant_stats: List[Dict[str, Any]] = []
    total_batches = 0
    total_abnormal = 0

    for data in per_tenant.values():
        abnormal = (
            data["low_stock"] + data["out_of_stock"] + data["near_expiry"] + data["expired"]
        )
        data["abnormal_batches"] = abnormal
        data["abnormal_ratio"] = (
            round(abnormal / data["total_batches"], 4) if data["total_batches"] else 0.0
        )
        tenant_stats.append(data)
        total_batches += data["total_batches"]
        total_abnormal += abnormal

    overall_ratio = round(total_abnormal / total_batches, 4) if total_batches else 0.0

    # 标记异常比例较高的企业
    tenant_stats.sort(key=lambda x: x["abnormal_ratio"], reverse=True)
    for t in tenant_stats:
        if t["abnormal_ratio"] >= 0.3:
            t["risk_level"] = "high"
        elif t["abnormal_ratio"] >= 0.1:
            t["risk_level"] = "medium"
        else:
            t["risk_level"] = "low"

    return {
        "summary": {
            "total_batches": total_batches,
            "total_abnormal_batches": total_abnormal,
            "abnormal_ratio": overall_ratio,
        },
        "by_tenant": tenant_stats,
    }


def _circulation_compliance(start: datetime, end: datetime, region: str | None) -> Dict[str, Any]:
    """
    流通合规分析：
    - 统计各运输状态记录数量
    - 简单评估是否存在延迟上报：SHIPPED -> IN_TRANSIT > 24h 视作延迟
    - 生成按天的时效性对比数据（达标率）
    """
    records_query = CirculationRecord.query.filter(
        CirculationRecord.timestamp >= start, CirculationRecord.timestamp <= end
    ).join(Order, CirculationRecord.order_id == Order.id)

    if region:
        like = f"%{region}%"
        records_query = records_query.join(
            Tenant, Order.buyer_tenant_id == Tenant.id
        ).filter(Tenant.address.ilike(like))

    records: List[CirculationRecord] = records_query.all()

    status_counts = defaultdict(int)
    for r in records:
        status_counts[r.transport_status] += 1

    # 以 tracking_number 分组，评估是否存在延迟上报
    grouped: Dict[str, List[CirculationRecord]] = defaultdict(list)
    for r in records:
        grouped[r.tracking_number].append(r)

    delayed_count = 0
    total_shipments = 0
    daily_total = defaultdict(int)
    daily_on_time = defaultdict(int)

    for tracking, recs in grouped.items():
        recs.sort(key=lambda x: x.timestamp or datetime.min)
        first_shipped = next(
            (r.timestamp for r in recs if r.transport_status == "SHIPPED"), None
        )
        first_in_transit = next(
            (r.timestamp for r in recs if r.transport_status == "IN_TRANSIT"), None
        )
        if not first_shipped:
            continue
        total_shipments += 1
        shipped_day = (first_shipped or start).date().isoformat()
        daily_total[shipped_day] += 1

        if first_in_transit:
            delta = first_in_transit - first_shipped
            if delta > timedelta(hours=24):
                delayed_count += 1
            else:
                daily_on_time[shipped_day] += 1
        else:
            # 一直没有 IN_TRANSIT，视作延迟
            delayed_count += 1

    delayed_ratio = (
        round(delayed_count / total_shipments, 4) if total_shipments else 0.0
    )

    # 构造时效性趋势（按天）列表
    trend: List[Dict[str, Any]] = []
    all_days = sorted(daily_total.keys())
    for day in all_days:
        total = daily_total[day]
        on_time = daily_on_time.get(day, 0)
        trend.append(
            {
                "date": day,
                "total_shipments": total,
                "on_time": on_time,
                "on_time_ratio": round(on_time / total, 4) if total else 0.0,
            }
        )

    return {
        "summary": {
            "total_records": len(records),
            "status_counts": dict(status_counts),
            "total_shipments": total_shipments,
            "delayed_shipments": delayed_count,
            "delayed_ratio": delayed_ratio,
        },
        "timeliness_trend": trend,
    }


def _enterprise_compliance(
    start: datetime, end: datetime, region: str | None
) -> Dict[str, Any]:
    """
    企业合规分析：
    - 从租户维度统计库存异常 + 流通延迟次数
    - 计算简单的合规率，并标记违规次数较多的企业
    """
    # 库存异常统计（沿用 _inventory_compliance 的规则）
    inv_data = _inventory_compliance(start, end, region)
    per_tenant = {t["tenant_id"]: t for t in inv_data["by_tenant"]}

    # 流通延迟统计（基于 _circulation_compliance 的分组逻辑）
    records_query = CirculationRecord.query.filter(
        CirculationRecord.timestamp >= start, CirculationRecord.timestamp <= end
    ).join(Order, CirculationRecord.order_id == Order.id)

    if region:
        like = f"%{region}%"
        records_query = records_query.join(
            Tenant, Order.buyer_tenant_id == Tenant.id
        ).filter(Tenant.address.ilike(like))

    records: List[CirculationRecord] = records_query.all()
    grouped_by_order: Dict[int, List[CirculationRecord]] = defaultdict(list)
    for r in records:
        if r.order_id:
            grouped_by_order[r.order_id].append(r)

    tenant_delay_counts: Dict[int, int] = defaultdict(int)

    for order_id, recs in grouped_by_order.items():
        order = Order.query.get(order_id)
        if not order:
            continue
        recs.sort(key=lambda x: x.timestamp or datetime.min)
        first_shipped = next(
            (r.timestamp for r in recs if r.transport_status == "SHIPPED"), None
        )
        first_in_transit = next(
            (r.timestamp for r in recs if r.transport_status == "IN_TRANSIT"), None
        )
        if not first_shipped:
            continue
        if not first_in_transit or first_in_transit - first_shipped > timedelta(hours=24):
            # 按买方和卖方同时记一次违规
            if order.buyer_tenant_id:
                tenant_delay_counts[order.buyer_tenant_id] += 1
            if order.supplier_tenant_id:
                tenant_delay_counts[order.supplier_tenant_id] += 1

    enterprise_rows: List[Dict[str, Any]] = []
    for tenant in Tenant.query.all():
        if region and region not in (tenant.address or ""):
            continue
        inv_stats = per_tenant.get(tenant.id, None)
        inventory_violations = (
            inv_stats["abnormal_batches"] if inv_stats else 0
        )
        circulation_violations = tenant_delay_counts.get(tenant.id, 0)
        total_violations = inventory_violations + circulation_violations
        total_checks = (
            (inv_stats["total_batches"] if inv_stats else 0)
            + max(circulation_violations, 1)
        )
        compliance_rate = (
            round(1 - (total_violations / total_checks), 4)
            if total_checks
            else 1.0
        )

        if total_violations >= 10:
            risk_level = "high"
        elif total_violations >= 3:
            risk_level = "medium"
        else:
            risk_level = "low"

        enterprise_rows.append(
            {
                "tenant_id": tenant.id,
                "tenant_name": tenant.name,
                "type": tenant.type,
                "inventory_violations": inventory_violations,
                "circulation_violations": circulation_violations,
                "total_violations": total_violations,
                "compliance_rate": compliance_rate,
                "risk_level": risk_level,
            }
        )

    enterprise_rows.sort(
        key=lambda x: (x["risk_level"] != "high", -x["total_violations"])
    )

    return {
        "items": enterprise_rows,
    }


def _build_report_payload(start: datetime, end: datetime, region: str | None) -> Dict[str, Any]:
    """汇总三大类合规分析结果，用于前端预览与导出。"""
    inventory = _inventory_compliance(start, end, region)
    circulation = _circulation_compliance(start, end, region)
    enterprise = _enterprise_compliance(start, end, region)
    return {
        "range": {
            "start": start.isoformat(),
            "end": end.isoformat(),
            "region": region,
        },
        "inventory": inventory,
        "circulation": circulation,
        "enterprise": enterprise,
    }


@bp.route("/report/preview", methods=["GET"])
@jwt_required()
def preview_report():
    """
    监管用户在浏览器内查看合规分析报告（JSON 数据），
    用于前端图表与表格展示。
    """
    user = _get_current_user()
    if not _require_regulator(user):
        return jsonify({"msg": "仅监管用户可访问合规分析报告"}), 403

    start, end = _parse_range()
    region = (request.args.get("region") or "").strip() or None

    payload = _build_report_payload(start, end, region)
    return jsonify({"msg": "ok", "data": payload})


def _export_pdf(payload: Dict[str, Any]) -> BytesIO:
    if not canvas or not A4 or not getSampleStyleSheet:
        raise RuntimeError("PDF 导出依赖 reportlab，请先在后端安装该依赖")

    # Use Latin font only; all labels below are English to avoid CJK garbling
    font_name = "Helvetica"

    buffer = BytesIO()
    c = canvas.Canvas(buffer, pagesize=A4)
    width, height = A4

    styles = getSampleStyleSheet()
    textobject = c.beginText(40, height - 50)
    textobject.setFont(font_name, 16)
    textobject.textLine("Drug Circulation Compliance Analysis Report")

    textobject.setFont(font_name, 11)
    r = payload["range"]
    textobject.textLine(
        f"Time range: {r['start'][:19]} to {r['end'][:19]}   Region: {r['region'] or 'All'}"
    )
    textobject.textLine("")

    # Inventory compliance summary
    inv = payload["inventory"]["summary"]
    textobject.setFont(font_name, 13)
    textobject.textLine("1. Inventory Compliance Analysis")
    textobject.setFont(font_name, 11)
    textobject.textLine(
        f"- Total batches: {inv['total_batches']}  Abnormal batches: {inv['total_abnormal_batches']}  "
        f"Abnormal ratio: {inv['abnormal_ratio'] * 100:.2f}%"
    )
    textobject.textLine("")

    # Circulation compliance summary
    circ = payload["circulation"]["summary"]
    textobject.setFont(font_name, 13)
    textobject.textLine("2. Circulation Compliance Analysis")
    textobject.setFont(font_name, 11)
    textobject.textLine(
        f"- Total records: {circ['total_records']}  Shipments: {circ['total_shipments']}  "
        f"Delayed shipments: {circ['delayed_shipments']}  Delay ratio: {circ['delayed_ratio'] * 100:.2f}%"
    )
    textobject.textLine("")

    # Enterprise compliance summary
    ent_items = payload["enterprise"]["items"]
    high_risk = [e for e in ent_items if e["risk_level"] == "high"][:10]
    textobject.setFont(font_name, 13)
    textobject.textLine("3. Enterprise Compliance Analysis")
    textobject.setFont(font_name, 11)
    textobject.textLine(f"- Enterprises in scope: {len(ent_items)}")
    if high_risk:
        textobject.textLine("- High-risk enterprises (top 10 by violations):")
        for e in high_risk:
            line = (
                f"  · {e['tenant_name']} (Type: {e['type']}, "
                f"Total violations: {e['total_violations']}, "
                f"Compliance rate: {e['compliance_rate'] * 100:.2f}%)"
            )
            textobject.textLine(line)
    else:
        textobject.textLine("- No high-risk enterprises in current range.")

    c.drawText(textobject)
    c.showPage()
    c.save()
    buffer.seek(0)
    return buffer


def _export_excel(payload: Dict[str, Any]) -> BytesIO:
    if not Workbook:
        raise RuntimeError("Excel 导出依赖 openpyxl，请先在后端安装该依赖")

    wb = Workbook()
    # 库存合规
    ws_inv = wb.active
    ws_inv.title = "库存合规"
    ws_inv.append(
        [
            "企业ID",
            "企业名称",
            "总批次",
            "低库存",
            "缺货",
            "近效期",
            "过期",
            "异常批次",
            "异常比例",
            "风险等级",
        ]
    )
    for t in payload["inventory"]["by_tenant"]:
        ws_inv.append(
            [
                t["tenant_id"],
                t["tenant_name"],
                t["total_batches"],
                t["low_stock"],
                t["out_of_stock"],
                t["near_expiry"],
                t["expired"],
                t["abnormal_batches"],
                t["abnormal_ratio"],
                t["risk_level"],
            ]
        )

    # 流通时效
    ws_circ = wb.create_sheet("流通时效")
    ws_circ.append(["日期", "发运批次", "按时批次", "按时比例"])
    for row in payload["circulation"]["timeliness_trend"]:
        ws_circ.append(
            [
                row["date"],
                row["total_shipments"],
                row["on_time"],
                row["on_time_ratio"],
            ]
        )

    # 企业合规
    ws_ent = wb.create_sheet("企业合规")
    ws_ent.append(
        [
            "企业ID",
            "企业名称",
            "类型",
            "库存违规",
            "流通违规",
            "总违规",
            "合规率",
            "风险等级",
        ]
    )
    for e in payload["enterprise"]["items"]:
        ws_ent.append(
            [
                e["tenant_id"],
                e["tenant_name"],
                e["type"],
                e["inventory_violations"],
                e["circulation_violations"],
                e["total_violations"],
                e["compliance_rate"],
                e["risk_level"],
            ]
        )

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer


@bp.route("/report/export", methods=["GET"])
@jwt_required()
def export_report():
    """
    导出合规分析报告为 PDF 或 Excel。
    - format=pdf: application/pdf
    - format=excel: application/vnd.openxmlformats-officedocument.spreadsheetml.sheet

    所有下载操作写入审计日志。
    """
    user = _get_current_user()
    if not _require_regulator(user):
        return jsonify({"msg": "仅监管用户可导出合规分析报告"}), 403

    fmt = (request.args.get("format") or "pdf").lower()
    if fmt not in {"pdf", "excel"}:
        return jsonify({"msg": "format 必须是 pdf 或 excel"}), 400

    start, end = _parse_range()
    region = (request.args.get("region") or "").strip() or None

    # 构造报告数据
    payload = _build_report_payload(start, end, region)

    # 记录下载行为到审计日志
    try:
        record_admin_action(
            user,
            "export_compliance_report",
            resource_type="compliance_report",
            resource_id=None,
            details={
                "format": fmt,
                "start": payload["range"]["start"],
                "end": payload["range"]["end"],
                "region": region,
            },
            commit=True,
        )
    except Exception as exc:  # pragma: no cover - 审计失败不影响主流程
        current_app.logger.exception("Failed to record compliance report export: %s", exc)

    filename = f"compliance-report-{start.date()}_{end.date()}.{ 'pdf' if fmt == 'pdf' else 'xlsx' }"

    try:
        if fmt == "pdf":
            buffer = _export_pdf(payload)
            return send_file(
                buffer,
                mimetype="application/pdf",
                as_attachment=True,
                download_name=filename,
            )
        else:
            buffer = _export_excel(payload)
            return send_file(
                buffer,
                mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                as_attachment=True,
                download_name=filename,
            )
    except RuntimeError as exc:
        return jsonify({"msg": str(exc)}), 500
    except Exception as exc:  # pragma: no cover
        current_app.logger.exception("Failed to export compliance report: %s", exc)
        return jsonify({"msg": "导出报告失败，请稍后重试"}), 500


